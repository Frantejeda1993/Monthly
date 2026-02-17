import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import numpy as np
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Sportech IB · Dashboard",
    page_icon="🏍️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Custom CSS ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@400;500&display=swap');

:root {
    --bg:       #0a0c10;
    --surface:  #12151c;
    --card:     #1a1f2b;
    --border:   #252d3d;
    --accent:   #e8ff00;
    --accent2:  #00e5ff;
    --danger:   #ff4757;
    --success:  #2ed573;
    --muted:    #5a6378;
    --text:     #d4dbe8;
    --text-dim: #8896ab;
}

html, body, [data-testid="stAppViewContainer"] {
    background-color: var(--bg) !important;
    color: var(--text);
    font-family: 'Syne', sans-serif;
}

[data-testid="stSidebar"] {
    background-color: var(--surface) !important;
    border-right: 1px solid var(--border);
}

[data-testid="stSidebar"] .block-container { padding-top: 2rem; }

h1, h2, h3 { font-family: 'Syne', sans-serif; font-weight: 800; }

/* Metric cards */
.metric-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 20px 24px;
    margin-bottom: 12px;
    transition: border-color 0.2s;
}
.metric-card:hover { border-color: var(--accent); }
.metric-label {
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: var(--text-dim);
    margin-bottom: 6px;
}
.metric-value {
    font-family: 'DM Mono', monospace;
    font-size: 28px;
    font-weight: 500;
    color: var(--text);
    line-height: 1;
}
.metric-delta {
    font-family: 'DM Mono', monospace;
    font-size: 12px;
    margin-top: 6px;
}
.delta-up   { color: var(--success); }
.delta-down { color: var(--danger); }
.delta-neu  { color: var(--text-dim); }

/* Section headers */
.section-header {
    font-size: 13px;
    font-weight: 700;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    color: var(--accent);
    padding: 12px 0 8px;
    border-bottom: 1px solid var(--border);
    margin-bottom: 16px;
}

/* Tag pill */
.tag {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 0.08em;
    text-transform: uppercase;
}
.tag-2w  { background: #1a2a4a; color: #7eb8ff; border: 1px solid #2a4a7a; }
.tag-ft  { background: #1a3a1a; color: #5de85d; border: 1px solid #2a5a2a; }
.tag-ot  { background: #3a1a3a; color: #e87de8; border: 1px solid #5a2a5a; }
.tag-tot { background: #2a2a1a; color: var(--accent);  border: 1px solid #4a4a2a; }

/* Sidebar nav */
.nav-item {
    padding: 10px 14px;
    border-radius: 8px;
    margin: 3px 0;
    cursor: pointer;
    font-weight: 600;
    font-size: 13px;
    transition: all 0.15s;
}

/* Plotly chart bg */
.js-plotly-plot .plotly { background: transparent !important; }

/* Streamlit overrides */
div[data-testid="stMetric"] {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 16px 20px;
}
div[data-testid="stMetricValue"] {
    font-family: 'DM Mono', monospace !important;
    font-size: 22px !important;
}
[data-testid="stSelectbox"] label,
[data-testid="stRadio"] label { color: var(--text-dim) !important; font-size: 12px !important; }

.stDataFrame { border-radius: 10px; overflow: hidden; }
thead tr th { background: var(--card) !important; }

hr { border-color: var(--border); }
</style>
""", unsafe_allow_html=True)

# ── Data loading ────────────────────────────────────────────────────────────────
@st.cache_data
def load_data(uploaded_file=None):
    """Load and process all data from the Excel file."""
    if uploaded_file is not None:
        file_source = BytesIO(uploaded_file.read())
    else:
        return None

    from openpyxl import load_workbook

    # ── RAW sheets ──
    wb_raw  = load_workbook(file_source, data_only=True)
    file_source.seek(0)

    # --- Ventas ---
    df_ventas = pd.read_excel(file_source, sheet_name='INPUT (Mensual) Ventas')
    file_source.seek(0)

    # --- Familias ---
    df_familias = pd.read_excel(file_source, sheet_name='INPUT (Anual) Familias')
    file_source.seek(0)

    # --- Budget ---
    df_budget_raw = pd.read_excel(file_source, sheet_name='INPUT (Anual) Budget')
    file_source.seek(0)

    # --- Margin LY ---
    df_margin_ly = pd.read_excel(file_source, sheet_name='INPUT (Anual) MARGIN LY',
                                  header=[0, 1])
    file_source.seek(0)

    # ── Compute monthly sales data ──
    df_ventas['Margen_Euros'] = (
        df_ventas['Importe Neto'] * df_ventas['CR3: % Margen s/Venta'] / 100
    )
    df_ventas_full = df_ventas.merge(
        df_familias[['Nombre', 'Familia', 'Columna1']],
        left_on='Clave 1', right_on='Familia', how='left'
    )

    # Per brand monthly
    brand_monthly = df_ventas_full.groupby(
        ['Nombre', 'Mes Factura']
    ).agg(Revenue=('Importe Neto','sum'),
          Margen_Euros=('Margen_Euros','sum')).reset_index()
    brand_monthly['Margen_Pct'] = np.where(
        brand_monthly['Revenue'] != 0,
        brand_monthly['Margen_Euros'] / brand_monthly['Revenue'],
        0
    )

    # Per vertical monthly
    vertical_monthly = df_ventas_full.groupby(
        ['Columna1', 'Mes Factura']
    ).agg(Revenue=('Importe Neto','sum'),
          Margen_Euros=('Margen_Euros','sum')).reset_index()
    vertical_monthly['Margen_Pct'] = np.where(
        vertical_monthly['Revenue'] != 0,
        vertical_monthly['Margen_Euros'] / vertical_monthly['Revenue'],
        0
    )

    # ── MARGINS table from openpyxl (static data like Stock, Budget, LY) ──
    ws_m = wb_raw['MARGINS']
    margins_rows = []
    header_rows = {4:'2 WHEELS', 39:'VARIOS', 49:'FREE TIME', 62:'OUTDOOR TECH', 72:'SIN CLASIFICAR'}
    vertical_map = {
        (5, 38): '2 WHEELS',
        (39, 48): 'VARIOS',
        (50, 61): 'FREE TIME',
        (63, 71): 'OUTDOOR TECH',
        (72, 73): 'SIN CLASIFICAR',
    }

    month_cols = {
        'Enero':      (16, 18),   # P=Revenue, R=Margen€
        'Febrero':    (19, 21),
        'Marzo':      (22, 24),
        'Abril':      (25, 27),
        'Mayo':       (28, 30),
        'Junio':      (31, 33),
        'Julio':      (34, 36),
        'Agosto':     (37, 39),
        'Septiembre': (40, 42),
        'Octubre':    (43, 45),
        'Noviembre':  (46, 48),
        'Diciembre':  (49, 51),
    }

    skip_rows = {3, 4, 39, 49, 62, 72}  # totals/headers rows
    totals_rows = {3, 5, 40, 50, 63, 73}

    for row in ws_m.iter_rows(min_row=3, max_row=73, max_col=51):
        rn = row[0].row
        marca = row[0].value
        if not marca or marca in ('Marca', '') or isinstance(row[1].value, str) and row[1].value == 'Stock':
            continue

        def safe(v):
            if isinstance(v, (int, float)): return v
            return 0

        vertical = '—'
        for (r1, r2), v in vertical_map.items():
            if r1 <= rn <= r2:
                vertical = v
                break

        rec = {
            'row':         rn,
            'Marca':       str(marca).strip(),
            'Vertical':    vertical,
            'Stock':       safe(row[1].value),
            'Budget_Rev':  safe(row[2].value),
            'Budget_Mg%':  safe(row[3].value),
            'Budget_MgEur':safe(row[4].value),
            'Acum_Rev':    safe(row[5].value),
            'Acum_Mg%':    safe(row[6].value),
            'Acum_MgEur':  safe(row[7].value),
            'LY_Rev':      safe(row[9].value),
            'LY_MgEur':    safe(row[10].value),
            'LY_Mg%':      safe(row[11].value),
        }
        for mes, (rc, rm) in month_cols.items():
            rec[f'{mes}_Rev']    = safe(row[rc-1].value)
            rec[f'{mes}_MgEur']  = safe(row[rm-1].value)
        margins_rows.append(rec)

    df_margins = pd.DataFrame(margins_rows)

    # ── Budget monthly totals ──
    month_budget_cols = {
        1:'Venta Enero',2:'Venta Febrero',3:'Venta Marzo',4:'Venta Abril',
        5:'Venta Mayo',6:'Venta Junio',7:'Venta Julio',8:'Venta Agosto',
        9:'Venta Septiembre',10:'Venta Octubre',11:'Venta Noviembre',12:'Venta Diciembre'
    }
    budget_monthly = {}
    for m, col in month_budget_cols.items():
        budget_monthly[m] = df_budget_raw[col].sum() if col in df_budget_raw.columns else 0

    return {
        'ventas':          df_ventas,
        'ventas_full':     df_ventas_full,
        'familias':        df_familias,
        'budget_raw':      df_budget_raw,
        'margins':         df_margins,
        'brand_monthly':   brand_monthly,
        'vertical_monthly':vertical_monthly,
        'budget_monthly':  budget_monthly,
    }


# ── Chart helpers ───────────────────────────────────────────────────────────────
CHART_LAYOUT = dict(
    paper_bgcolor='rgba(0,0,0,0)',
    plot_bgcolor='rgba(0,0,0,0)',
    font_family='Syne',
    font_color='#d4dbe8',
    margin=dict(l=0, r=0, t=32, b=0),
    legend=dict(
        bgcolor='rgba(26,31,43,0.8)',
        bordercolor='#252d3d',
        borderwidth=1,
        font_size=11,
    ),
    xaxis=dict(gridcolor='#252d3d', linecolor='#252d3d', zerolinecolor='#252d3d'),
    yaxis=dict(gridcolor='#252d3d', linecolor='#252d3d', zerolinecolor='#252d3d'),
)

VERTICAL_COLORS = {
    '2 WHEELS':    '#7eb8ff',
    'FREE TIME':   '#5de85d',
    'OUTDOOR TECH':'#e87de8',
    'VARIOS':      '#f5a623',
    'TOTAL':       '#e8ff00',
    '—':           '#8896ab',
}

def fmt_eur(v, decimals=0):
    if abs(v) >= 1_000_000:
        return f"€{v/1_000_000:.1f}M"
    elif abs(v) >= 1_000:
        return f"€{v/1_000:.0f}K"
    return f"€{v:,.{decimals}f}"

def fmt_pct(v):
    return f"{v*100:.1f}%"

def metric_card(label, value, delta=None, delta_label="vs LY"):
    delta_html = ""
    if delta is not None:
        if isinstance(delta, float) and abs(delta) < 10:
            dval = f"{delta*100:+.1f}pp" if "%" in delta_label else f"{delta:+.1f}%"
        else:
            dval = fmt_eur(delta) if delta > 100 else f"{delta:+.0f}"
        cls = "delta-up" if (delta if isinstance(delta, (int,float)) else 0) >= 0 else "delta-down"
        delta_html = f'<div class="metric-delta {cls}">{dval} {delta_label}</div>'
    return f"""
    <div class="metric-card">
        <div class="metric-label">{label}</div>
        <div class="metric-value">{value}</div>
        {delta_html}
    </div>"""


# ── Sidebar ─────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 🏍️ **Sportech IB**")
    st.markdown('<div style="color:#5a6378;font-size:11px;margin-bottom:20px;">Monthly Performance Dashboard</div>', unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "Cargar Excel",
        type=["xlsx"],
        help="Sube el TemplateMonthly para actualizar los datos"
    )

    st.markdown("---")
    page = st.radio(
        "Sección",
        ["📊 Overview · RECAP", "📈 MARGINS · Marcas", "🏍️ Vertical 2 Wheels",
         "🌲 Vertical Free Time", "📡 Vertical Outdoor Tech"],
        label_visibility="collapsed"
    )
    st.markdown("---")
    st.markdown('<div style="color:#5a6378;font-size:10px;">Datos: Enero 2026 · Sportech IB</div>', unsafe_allow_html=True)


# ── Load data ───────────────────────────────────────────────────────────────────
if uploaded is None:
    st.markdown("""
    <div style="display:flex;flex-direction:column;align-items:center;justify-content:center;
                height:60vh;gap:16px;">
        <div style="font-size:64px;">🏍️</div>
        <div style="font-size:28px;font-weight:800;color:#e8ff00;">Sportech IB Dashboard</div>
        <div style="color:#5a6378;font-size:14px;">Sube el archivo <strong>TemplateMonthly.xlsx</strong> desde el panel lateral</div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

data = load_data(uploaded)
if data is None:
    st.error("No se pudo cargar el archivo.")
    st.stop()

df_margins      = data['margins']
brand_monthly   = data['brand_monthly']
vert_monthly    = data['vertical_monthly']
budget_monthly  = data['budget_monthly']
df_ventas       = data['ventas']
df_ventas_full  = data['ventas_full']
df_budget_raw   = data['budget_raw']

# Computed totals for current month
MONTHS_ES = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
             7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}

avail_months = sorted(brand_monthly['Mes Factura'].unique())
current_month = avail_months[-1] if avail_months else 1
current_month_name = MONTHS_ES.get(current_month, str(current_month))

total_rev   = df_ventas_full['Importe Neto'].sum()
total_mg_eur= df_ventas_full['Margen_Euros'].sum()
total_mg_pct= total_mg_eur / total_rev if total_rev else 0
total_stock = df_margins['Stock'].sum()
budget_enero= budget_monthly.get(1, 0)


# ════════════════════════════════════════════════════════════════════════════════
# PAGE 1: OVERVIEW · RECAP
# ════════════════════════════════════════════════════════════════════════════════
if page == "📊 Overview · RECAP":

    st.markdown(f"""
    <div style="display:flex;align-items:baseline;gap:12px;margin-bottom:8px;">
        <h1 style="margin:0;font-size:32px;">RECAP</h1>
        <span style="color:#5a6378;font-size:14px;">Resumen Global · {current_month_name} 2026</span>
    </div>""", unsafe_allow_html=True)

    # ── KPI row ──
    c1, c2, c3, c4, c5 = st.columns(5)
    ly_total = df_margins.loc[df_margins['Marca']=='TOTAL','LY_Rev'].values
    ly_rev   = ly_total[0] if len(ly_total) > 0 else 0
    ly_mg    = df_margins.loc[df_margins['Marca']=='TOTAL','LY_MgEur'].values
    ly_mg_eur= ly_mg[0] if len(ly_mg) > 0 else 0
    ly_mg_pct= ly_mg_eur / ly_rev if ly_rev else 0
    budget_tot= df_margins.loc[df_margins['Marca']=='TOTAL','Budget_Rev'].values
    budget_rev= float(budget_tot[0]) if len(budget_tot) > 0 else 0

    with c1:
        st.metric("Revenue Enero", fmt_eur(total_rev),
                  f"{(total_rev/ly_rev*12 - 1)*100:+.1f}% vs LY (anualiz.)" if ly_rev else "—")
    with c2:
        st.metric("Margen €", fmt_eur(total_mg_eur),
                  f"{(total_mg_pct - ly_mg_pct)*100:+.1f}pp vs LY" if ly_rev else "—")
    with c3:
        st.metric("Margen %", fmt_pct(total_mg_pct),
                  f"Budget: {fmt_pct(df_margins.loc[df_margins['Marca']=='TOTAL','Budget_Mg%'].values[0] if len(df_margins.loc[df_margins['Marca']=='TOTAL']) else 0)}")
    with c4:
        st.metric("Stock", fmt_eur(total_stock))
    with c5:
        bud_ene = budget_monthly.get(1, 0)
        cumpl = total_rev / bud_ene if bud_ene else 0
        st.metric("Cumpl. Budget", fmt_pct(cumpl),
                  f"Budget: {fmt_eur(bud_ene)}")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Revenue by vertical (bar) + Margin % (scatter) ──
    col_left, col_right = st.columns([3, 2])

    with col_left:
        st.markdown('<div class="section-header">Revenue por Vertical</div>', unsafe_allow_html=True)

        verticals_summary = vert_monthly.groupby('Columna1').agg(
            Revenue=('Revenue','sum'), Margen_Euros=('Margen_Euros','sum')
        ).reset_index().dropna()
        verticals_summary['Margen_Pct'] = verticals_summary['Margen_Euros'] / verticals_summary['Revenue']

        # Add budget per vertical using margins data
        v_budget = df_margins[df_margins['Marca'].isin(['2 WHEELS','FREE TIME','OUTDOOR TECH'])][['Marca','Budget_Rev']].copy()
        v_budget.columns = ['Columna1','Budget']
        v_budget['Columna1'] = v_budget['Columna1'].str.upper().str.strip()
        verticals_summary['Columna1_upper'] = verticals_summary['Columna1'].str.upper().str.strip()

        fig = go.Figure()
        for _, row in verticals_summary.iterrows():
            vname = str(row['Columna1'])
            color = VERTICAL_COLORS.get(vname.upper(), '#7eb8ff')
            bud_match = v_budget[v_budget['Columna1']==vname.upper()]
            bud_val = float(bud_match['Budget'].iloc[0]) if len(bud_match) else 0
            monthly_budget = bud_val / 12

            fig.add_trace(go.Bar(
                name=vname,
                x=[vname],
                y=[row['Revenue']],
                marker_color=color,
                marker_line_width=0,
                text=[fmt_eur(row['Revenue'])],
                textposition='outside',
                textfont=dict(size=11, color=color),
                hovertemplate=f"<b>{vname}</b><br>Revenue: €%{{y:,.0f}}<br>Margen: {fmt_pct(row['Margen_Pct'])}<extra></extra>",
            ))
            if monthly_budget > 0:
                fig.add_shape(type='line',
                    x0=-0.4+list(verticals_summary['Columna1']).index(vname),
                    x1=0.4+list(verticals_summary['Columna1']).index(vname),
                    y0=monthly_budget, y1=monthly_budget,
                    line=dict(color='#e8ff00', width=2, dash='dot'),
                )

        fig.update_layout(**CHART_LAYOUT, height=320, showlegend=False,
                          yaxis_title=None, xaxis_title=None,
                          bargap=0.3)
        fig.add_annotation(x=0.98, y=0.96, xref='paper', yref='paper',
                           text="— Budget mensual", font=dict(size=10, color='#e8ff00'),
                           showarrow=False, align='right')
        st.plotly_chart(fig, use_container_width=True)

    with col_right:
        st.markdown('<div class="section-header">Margen % por Vertical</div>', unsafe_allow_html=True)

        fig2 = go.Figure()
        for _, row in verticals_summary.iterrows():
            vname = str(row['Columna1'])
            color = VERTICAL_COLORS.get(vname.upper(), '#7eb8ff')
            fig2.add_trace(go.Bar(
                name=vname,
                x=[fmt_pct(row['Margen_Pct'])],
                y=[vname],
                orientation='h',
                marker_color=color,
                marker_line_width=0,
                text=[fmt_pct(row['Margen_Pct'])],
                textposition='outside',
                textfont=dict(size=11, color=color),
                hovertemplate=f"<b>{vname}</b><br>Margen: {fmt_pct(row['Margen_Pct'])}<extra></extra>",
            ))

        fig2.update_layout(**CHART_LAYOUT, height=320, showlegend=False,
                           xaxis=dict(tickformat='.0%', gridcolor='#252d3d', linecolor='#252d3d'),
                           yaxis=dict(gridcolor='#252d3d', linecolor='#252d3d'),
                           bargap=0.35)
        st.plotly_chart(fig2, use_container_width=True)

    # ── Budget evolution ──
    st.markdown('<div class="section-header">Budget Mensual 2026 — Distribución</div>', unsafe_allow_html=True)
    months_list = list(MONTHS_ES.values())
    budget_vals = [budget_monthly.get(m, 0) for m in range(1, 13)]
    cum_budget  = np.cumsum(budget_vals)

    fig3 = make_subplots(specs=[[{"secondary_y": True}]])
    fig3.add_trace(go.Bar(
        x=months_list, y=budget_vals,
        marker_color='#252d3d',
        marker_line_color='#3d4b63', marker_line_width=1,
        name='Budget Mensual',
        hovertemplate='%{x}: €%{y:,.0f}<extra></extra>',
    ), secondary_y=False)
    # Mark enero actual
    fig3.add_trace(go.Bar(
        x=[months_list[current_month-1]],
        y=[total_rev],
        marker_color='#e8ff00',
        name='Ventas Real',
        hovertemplate='Real: €%{y:,.0f}<extra></extra>',
    ), secondary_y=False)
    fig3.add_trace(go.Scatter(
        x=months_list, y=cum_budget,
        mode='lines+markers',
        line=dict(color='#00e5ff', width=2),
        marker=dict(size=5, color='#00e5ff'),
        name='Acum. Budget',
        hovertemplate='%{x} Acum: €%{y:,.0f}<extra></extra>',
    ), secondary_y=True)

    fig3.update_layout(**CHART_LAYOUT, height=280, bargap=0.15,
                       legend=dict(orientation='h', y=1.1, x=0))
    fig3.update_yaxes(gridcolor='#252d3d', linecolor='#252d3d', secondary_y=False)
    fig3.update_yaxes(gridcolor='rgba(0,0,0,0)', linecolor='#252d3d', secondary_y=True)
    st.plotly_chart(fig3, use_container_width=True)

    # ── Top brands table ──
    st.markdown('<div class="section-header">Top Marcas · Enero 2026</div>', unsafe_allow_html=True)
    top_brands = brand_monthly[brand_monthly['Mes Factura']==current_month].copy()
    top_brands = top_brands[top_brands['Nombre'].notna() & (top_brands['Nombre'] != '0 - SIN CLASIFICAR')]
    top_brands = top_brands.sort_values('Revenue', ascending=False).head(15)

    # Merge vertical info
    top_brands = top_brands.merge(
        data['familias'][['Nombre','Columna1']].drop_duplicates(),
        on='Nombre', how='left'
    )

    def render_table(df):
        rows = ""
        for _, r in df.iterrows():
            v = str(r.get('Columna1','—') or '—').upper().strip()
            tag_cls = {'2 WHEELS':'tag-2w','FREE TIME':'tag-ft','OUTDOOR TECH':'tag-ot'}.get(v,'')
            tag_label = {'2 WHEELS':'2W','FREE TIME':'FT','OUTDOOR TECH':'OT'}.get(v,'—')
            mg_color = '#2ed573' if r['Margen_Pct'] > 0.2 else ('#ff4757' if r['Margen_Pct'] < 0 else '#d4dbe8')
            rev_pct = r['Revenue'] / total_rev * 100
            rows += f"""
            <tr style="border-bottom:1px solid #1a1f2b;">
                <td style="padding:8px 12px;font-weight:600;">{r['Nombre']}</td>
                <td style="padding:8px 12px;"><span class="tag {tag_cls}">{tag_label}</span></td>
                <td style="padding:8px 12px;font-family:'DM Mono',monospace;text-align:right;">€{r['Revenue']:>10,.0f}</td>
                <td style="padding:8px 12px;font-family:'DM Mono',monospace;text-align:right;">€{r['Margen_Euros']:>8,.0f}</td>
                <td style="padding:8px 12px;font-family:'DM Mono',monospace;text-align:right;color:{mg_color};">{r['Margen_Pct']*100:.1f}%</td>
                <td style="padding:8px 12px;">
                    <div style="background:#252d3d;border-radius:4px;height:6px;">
                        <div style="background:#e8ff00;border-radius:4px;height:6px;width:{min(rev_pct*3,100):.0f}%;"></div>
                    </div>
                </td>
            </tr>"""
        return f"""
        <table style="width:100%;border-collapse:collapse;font-size:13px;">
            <thead>
                <tr style="background:#12151c;border-bottom:1px solid #252d3d;">
                    <th style="padding:8px 12px;text-align:left;color:#5a6378;font-size:11px;letter-spacing:.1em;text-transform:uppercase;">Marca</th>
                    <th style="padding:8px 12px;text-align:left;color:#5a6378;font-size:11px;">Vert.</th>
                    <th style="padding:8px 12px;text-align:right;color:#5a6378;font-size:11px;">Revenue</th>
                    <th style="padding:8px 12px;text-align:right;color:#5a6378;font-size:11px;">Margen €</th>
                    <th style="padding:8px 12px;text-align:right;color:#5a6378;font-size:11px;">Mg%</th>
                    <th style="padding:8px 12px;text-align:left;color:#5a6378;font-size:11px;">Share</th>
                </tr>
            </thead>
            <tbody style="background:#1a1f2b;">{rows}</tbody>
        </table>"""

    st.markdown(render_table(top_brands), unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════════
# PAGE 2: MARGINS
# ════════════════════════════════════════════════════════════════════════════════
elif page == "📈 MARGINS · Marcas":

    st.markdown("""
    <div style="display:flex;align-items:baseline;gap:12px;margin-bottom:8px;">
        <h1 style="margin:0;font-size:32px;">MARGINS</h1>
        <span style="color:#5a6378;font-size:14px;">Detalle por Marca · Datos 2026</span>
    </div>""", unsafe_allow_html=True)

    # Filter controls
    fc1, fc2 = st.columns([2, 1])
    with fc1:
        vertical_filter = st.multiselect(
            "Filtrar por Vertical",
            options=['2 WHEELS', 'FREE TIME', 'OUTDOOR TECH', 'VARIOS'],
            default=['2 WHEELS', 'FREE TIME', 'OUTDOOR TECH'],
        )
    with fc2:
        sort_by = st.selectbox("Ordenar por", ["Budget Revenue", "Stock", "LY Revenue", "Margen% Budget"])

    df_fil = df_margins[
        df_margins['Vertical'].isin(vertical_filter) &
        ~df_margins['Marca'].isin(['TOTAL','2 WHEELS','FREE TIME','OUTDOOR TECH','VARIOS',
                                    'SIN CLASIFICAR','NUEVAS MARCAS','0 - SIN CLASIFICAR','—'])
    ].copy()

    sort_map = {"Budget Revenue":"Budget_Rev","Stock":"Stock","LY Revenue":"LY_Rev","Margen% Budget":"Budget_Mg%"}
    df_fil = df_fil.sort_values(sort_map[sort_by], ascending=False)

    # KPIs
    k1, k2, k3, k4 = st.columns(4)
    with k1: st.metric("Marcas activas", len(df_fil[df_fil['Budget_Rev'] > 0]))
    with k2: st.metric("Budget total", fmt_eur(df_fil['Budget_Rev'].sum()))
    with k3: st.metric("Stock total", fmt_eur(df_fil['Stock'].sum()))
    with k4: st.metric("Revenue LY total", fmt_eur(df_fil['LY_Rev'].sum()))

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Scatter: Budget vs LY ──
    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown('<div class="section-header">Budget vs Revenue LY por Marca</div>', unsafe_allow_html=True)
        scatter_df = df_fil[df_fil['Budget_Rev'] > 0].copy()
        scatter_df['color'] = scatter_df['Vertical'].map(VERTICAL_COLORS)

        fig_sc = px.scatter(
            scatter_df,
            x='LY_Rev', y='Budget_Rev',
            size='Stock',
            color='Vertical',
            color_discrete_map=VERTICAL_COLORS,
            hover_name='Marca',
            hover_data={'LY_Rev':':.0f','Budget_Rev':':.0f','Stock':':.0f','Vertical':False},
            size_max=40,
            labels={'LY_Rev':'Revenue LY (€)','Budget_Rev':'Budget 2026 (€)'},
        )
        # Diagonal reference line
        max_val = max(scatter_df['LY_Rev'].max(), scatter_df['Budget_Rev'].max())
        fig_sc.add_trace(go.Scatter(
            x=[0, max_val], y=[0, max_val],
            mode='lines', line=dict(color='#252d3d', dash='dot', width=1),
            showlegend=False, hoverinfo='skip'
        ))
        fig_sc.update_layout(**CHART_LAYOUT, height=340)
        st.plotly_chart(fig_sc, use_container_width=True)

    with col_b:
        st.markdown('<div class="section-header">Margen % Budget vs LY</div>', unsafe_allow_html=True)
        mg_df = df_fil[(df_fil['Budget_Mg%'] != 0) | (df_fil['LY_Mg%'] != 0)].copy()
        mg_df = mg_df[mg_df['Budget_Rev'] > 5000].head(20)

        fig_mg = go.Figure()
        colors = mg_df['Vertical'].map(VERTICAL_COLORS).tolist()
        fig_mg.add_trace(go.Bar(
            name='Margen% LY',
            x=mg_df['Marca'],
            y=mg_df['LY_Mg%'],
            marker_color='#252d3d',
            marker_line_color='#3d4b63',
            marker_line_width=1,
        ))
        fig_mg.add_trace(go.Bar(
            name='Margen% Budget',
            x=mg_df['Marca'],
            y=mg_df['Budget_Mg%'],
            marker_color=[VERTICAL_COLORS.get(v,'#7eb8ff') for v in mg_df['Vertical']],
            marker_opacity=0.8,
        ))
        fig_mg.update_layout(**CHART_LAYOUT, height=340, barmode='group', bargap=0.15,
                             yaxis=dict(tickformat='.0%', gridcolor='#252d3d', linecolor='#252d3d'),
                             xaxis=dict(tickangle=-35, gridcolor='#252d3d', linecolor='#252d3d'))
        st.plotly_chart(fig_mg, use_container_width=True)

    # ── Main margins table ──
    st.markdown('<div class="section-header">Tabla de Márgenes</div>', unsafe_allow_html=True)

    display_cols = {
        'Marca':'Marca','Vertical':'Vertical','Stock':'Stock',
        'Budget_Rev':'Budget Rev.','Budget_Mg%':'Budget Mg%',
        'LY_Rev':'LY Revenue','LY_Mg%':'LY Mg%','LY_MgEur':'LY Mg€'
    }
    tbl = df_fil[list(display_cols.keys())].copy()
    tbl.columns = list(display_cols.values())

    def style_margins(df):
        def color_mg(val):
            if isinstance(val, float):
                if val > 0.25: return 'color: #2ed573'
                if val < 0:    return 'color: #ff4757'
                if val > 0.15: return 'color: #d4dbe8'
            return 'color: #8896ab'

        styled = df.style\
            .format({
                'Stock':       '€{:,.0f}',
                'Budget Rev.': '€{:,.0f}',
                'Budget Mg%':  '{:.1%}',
                'LY Revenue':  '€{:,.0f}',
                'LY Mg%':      '{:.1%}',
                'LY Mg€':      '€{:,.0f}',
            })\
            .applymap(color_mg, subset=['Budget Mg%', 'LY Mg%'])\
            .set_properties(**{'background-color': '#1a1f2b', 'color': '#d4dbe8',
                               'border-color': '#252d3d', 'font-size': '12px'})
        return styled

    st.dataframe(style_margins(tbl), use_container_width=True, height=420)

    # ── Stock waterfall ──
    st.markdown('<div class="section-header">Stock por Marca (Top 20)</div>', unsafe_allow_html=True)
    stock_df = df_fil[df_fil['Stock'] > 0].nlargest(20, 'Stock')
    fig_stock = go.Figure(go.Bar(
        x=stock_df['Marca'],
        y=stock_df['Stock'],
        marker_color=[VERTICAL_COLORS.get(v, '#7eb8ff') for v in stock_df['Vertical']],
        text=[fmt_eur(v) for v in stock_df['Stock']],
        textposition='outside',
        textfont=dict(size=10),
    ))
    fig_stock.update_layout(**CHART_LAYOUT, height=280, bargap=0.25,
                            xaxis=dict(tickangle=-35, gridcolor='#252d3d', linecolor='#252d3d'))
    st.plotly_chart(fig_stock, use_container_width=True)


# ════════════════════════════════════════════════════════════════════════════════
# VERTICAL PAGES (shared renderer)
# ════════════════════════════════════════════════════════════════════════════════
else:
    # Determine which vertical
    vertical_config = {
        "🏍️ Vertical 2 Wheels":      {'key':'2 WHEELS',    'label':'2 Wheels',    'color':'#7eb8ff', 'icon':'🏍️'},
        "🌲 Vertical Free Time":      {'key':'FREE TIME',   'label':'Free Time',   'color':'#5de85d', 'icon':'🌲'},
        "📡 Vertical Outdoor Tech":   {'key':'OUTDOOR TECH','label':'Outdoor Tech','color':'#e87de8', 'icon':'📡'},
    }
    cfg = vertical_config[page]
    V_KEY   = cfg['key']
    V_LABEL = cfg['label']
    V_COLOR = cfg['color']
    V_ICON  = cfg['icon']

    st.markdown(f"""
    <div style="display:flex;align-items:baseline;gap:12px;margin-bottom:8px;">
        <h1 style="margin:0;font-size:32px;color:{V_COLOR};">{V_ICON} {V_LABEL}</h1>
        <span style="color:#5a6378;font-size:14px;">Vertical Dashboard · {current_month_name} 2026</span>
    </div>""", unsafe_allow_html=True)

    # Filter data for this vertical
    brands_v = df_margins[df_margins['Vertical']==V_KEY].copy()
    brands_v = brands_v[~brands_v['Marca'].isin([V_KEY,'NUEVAS MARCAS','Marca',''])]

    # Actual sales from ventas
    ventas_v = vert_monthly[vert_monthly['Columna1'].str.upper()==V_KEY].copy()
    brand_v_actual = df_ventas_full[
        df_ventas_full['Columna1'].str.upper()==V_KEY
    ].groupby('Nombre').agg(
        Revenue=('Importe Neto','sum'),
        Margen_Euros=('Margen_Euros','sum')
    ).reset_index()
    brand_v_actual['Margen_Pct'] = np.where(
        brand_v_actual['Revenue'] != 0,
        brand_v_actual['Margen_Euros'] / brand_v_actual['Revenue'], 0
    )

    rev_v  = brand_v_actual['Revenue'].sum()
    mg_v   = brand_v_actual['Margen_Euros'].sum()
    mgpct_v= mg_v / rev_v if rev_v else 0

    # Budget for this vertical
    v_row = df_margins[df_margins['Marca'].str.upper()==V_KEY]
    budget_v     = float(v_row['Budget_Rev'].iloc[0]) if len(v_row) else 0
    budget_mg_v  = float(v_row['Budget_Mg%'].iloc[0]) if len(v_row) else 0
    stock_v      = float(v_row['Stock'].iloc[0]) if len(v_row) else 0
    ly_rev_v     = float(v_row['LY_Rev'].iloc[0]) if len(v_row) else 0
    ly_mgeur_v   = float(v_row['LY_MgEur'].iloc[0]) if len(v_row) else 0
    ly_mgpct_v   = float(v_row['LY_Mg%'].iloc[0]) if len(v_row) else 0

    bud_monthly_v= budget_v / 12

    # ── KPI row ──
    k1, k2, k3, k4, k5 = st.columns(5)
    with k1:
        st.metric("Revenue Enero", fmt_eur(rev_v),
                  f"Budget: {fmt_eur(bud_monthly_v)}")
    with k2:
        cumpl_v = rev_v / bud_monthly_v if bud_monthly_v else 0
        st.metric("Cumpl. Budget", fmt_pct(cumpl_v),
                  f"{rev_v - bud_monthly_v:+,.0f}€ vs budget")
    with k3:
        st.metric("Margen €", fmt_eur(mg_v),
                  f"LY: {fmt_eur(ly_mgeur_v/12)}/mes")
    with k4:
        st.metric("Margen %", fmt_pct(mgpct_v),
                  f"{(mgpct_v - ly_mgpct_v)*100:+.1f}pp vs LY")
    with k5:
        st.metric("Stock", fmt_eur(stock_v))

    # Are we growing? HOW DOING vs BUDGET? LAST MONTH TREND
    st.markdown("<br>", unsafe_allow_html=True)
    ia1, ia2, ia3 = st.columns(3)
    ly_monthly = ly_rev_v / 12
    with ia1:
        growing = rev_v > ly_monthly
        indicator = "▲ GROWING" if growing else "▼ DECLINING"
        ind_color = "#2ed573" if growing else "#ff4757"
        pct_vs_ly = (rev_v / ly_monthly - 1)*100 if ly_monthly else 0
        st.markdown(f"""
        <div class="metric-card" style="border-color:{ind_color}40;">
            <div class="metric-label">Are we growing?</div>
            <div class="metric-value" style="font-size:20px;color:{ind_color};">{indicator}</div>
            <div class="metric-delta" style="color:{ind_color};">{pct_vs_ly:+.1f}% vs LY (mensualiz.)</div>
        </div>""", unsafe_allow_html=True)
    with ia2:
        doing_ok = rev_v >= bud_monthly_v * 0.9
        doing_label = "✓ ON TRACK" if doing_ok else "✗ BEHIND BUDGET"
        doing_color = "#2ed573" if doing_ok else "#ff4757"
        diff_eur = rev_v - bud_monthly_v
        st.markdown(f"""
        <div class="metric-card" style="border-color:{doing_color}40;">
            <div class="metric-label">How are we doing vs Budget?</div>
            <div class="metric-value" style="font-size:20px;color:{doing_color};">{doing_label}</div>
            <div class="metric-delta" style="color:{doing_color};">{diff_eur:+,.0f}€ vs budget mensual</div>
        </div>""", unsafe_allow_html=True)
    with ia3:
        stock_pct = stock_v / budget_v if budget_v else 0
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">% Stock vs Year Budget</div>
            <div class="metric-value" style="font-family:'DM Mono',monospace;">{fmt_pct(stock_pct)}</div>
            <div class="metric-delta delta-neu">Stock: {fmt_eur(stock_v)} | Budget anual: {fmt_eur(budget_v)}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Charts ──
    col_l, col_r = st.columns([3, 2])

    with col_l:
        st.markdown('<div class="section-header">Revenue Real vs Budget por Marca</div>', unsafe_allow_html=True)
        merged_v = brand_v_actual.merge(
            brands_v[['Marca','Budget_Rev','LY_Rev']].rename(columns={'Marca':'Nombre'}),
            on='Nombre', how='outer'
        ).fillna(0)
        merged_v = merged_v.sort_values('Revenue', ascending=False).head(15)

        fig_bv = go.Figure()
        fig_bv.add_trace(go.Bar(
            name='LY (anualiz./12)',
            x=merged_v['Nombre'],
            y=merged_v['LY_Rev'] / 12,
            marker_color='#252d3d',
            marker_line_color='#3d4b63', marker_line_width=1,
        ))
        fig_bv.add_trace(go.Bar(
            name='Revenue Enero',
            x=merged_v['Nombre'],
            y=merged_v['Revenue'],
            marker_color=V_COLOR,
            marker_opacity=0.85,
        ))
        fig_bv.add_trace(go.Scatter(
            name='Budget /mes',
            x=merged_v['Nombre'],
            y=merged_v['Budget_Rev'] / 12,
            mode='markers',
            marker=dict(symbol='diamond', size=9, color='#e8ff00',
                        line=dict(color='#0a0c10', width=1)),
        ))
        fig_bv.update_layout(**CHART_LAYOUT, height=340, barmode='group', bargap=0.2,
                             xaxis=dict(tickangle=-35, gridcolor='#252d3d', linecolor='#252d3d'),
                             legend=dict(orientation='h', y=1.1))
        st.plotly_chart(fig_bv, use_container_width=True)

    with col_r:
        st.markdown('<div class="section-header">Margen % Real por Marca</div>', unsafe_allow_html=True)
        mg_brand = brand_v_actual[brand_v_actual['Revenue'] > 0].sort_values('Revenue', ascending=False).head(12)
        colors_mg = ['#2ed573' if m > 0.2 else ('#ff4757' if m < 0 else V_COLOR) for m in mg_brand['Margen_Pct']]

        fig_mgb = go.Figure(go.Bar(
            x=mg_brand['Margen_Pct'],
            y=mg_brand['Nombre'],
            orientation='h',
            marker_color=colors_mg,
            text=[fmt_pct(m) for m in mg_brand['Margen_Pct']],
            textposition='outside',
            textfont=dict(size=10),
        ))
        # Budget margin line
        if budget_mg_v:
            fig_mgb.add_vline(x=budget_mg_v, line_color='#e8ff00', line_dash='dot', line_width=2,
                              annotation_text=f"Budget {fmt_pct(budget_mg_v)}", annotation_font_size=10)
        fig_mgb.update_layout(**CHART_LAYOUT, height=340, showlegend=False,
                              xaxis=dict(tickformat='.0%', gridcolor='#252d3d', linecolor='#252d3d'),
                              yaxis=dict(gridcolor='#252d3d', linecolor='#252d3d'))
        st.plotly_chart(fig_mgb, use_container_width=True)

    # ── Budget monthly distribution ──
    st.markdown('<div class="section-header">Distribución Budget Mensual 2026</div>', unsafe_allow_html=True)

    # Get budget by month for this vertical from budget table
    vert_brand_list = brands_v['Marca'].tolist()
    vert_budget_rows = df_budget_raw[df_budget_raw['Marca'].isin(vert_brand_list)]
    month_cols_budget = {
        'Enero':'Venta Enero','Febrero':'Venta Febrero','Marzo':'Venta Marzo',
        'Abril':'Venta Abril','Mayo':'Venta Mayo','Junio':'Venta Junio',
        'Julio':'Venta Julio','Agosto':'Venta Agosto','Septiembre':'Venta Septiembre',
        'Octubre':'Venta Octubre','Noviembre':'Venta Noviembre','Diciembre':'Venta Diciembre'
    }
    bud_by_month = []
    for mes, col in month_cols_budget.items():
        val = vert_budget_rows[col].sum() if col in vert_budget_rows.columns else 0
        bud_by_month.append({'Mes': mes, 'Budget': val})
    df_bm = pd.DataFrame(bud_by_month)

    fig_bm = go.Figure()
    fig_bm.add_trace(go.Bar(
        x=df_bm['Mes'],
        y=df_bm['Budget'],
        marker_color=V_COLOR,
        marker_opacity=0.3,
        name='Budget',
        hovertemplate='%{x}: €%{y:,.0f}<extra></extra>',
    ))
    # Actual for current month
    fig_bm.add_trace(go.Bar(
        x=[current_month_name],
        y=[rev_v],
        marker_color=V_COLOR,
        name='Real',
        hovertemplate=f'Real {current_month_name}: €%{{y:,.0f}}<extra></extra>',
    ))
    fig_bm.add_hline(y=bud_monthly_v, line_color='#e8ff00', line_dash='dot', line_width=1.5,
                     annotation_text=f"Media mensual {fmt_eur(bud_monthly_v)}", annotation_font_size=10)
    fig_bm.update_layout(**CHART_LAYOUT, height=260, bargap=0.2,
                         legend=dict(orientation='h', y=1.1))
    st.plotly_chart(fig_bm, use_container_width=True)

    # ── Detailed brands table ──
    st.markdown('<div class="section-header">Detalle de Marcas</div>', unsafe_allow_html=True)
    detail = brand_v_actual.merge(
        brands_v[['Marca','Budget_Rev','Budget_Mg%','Stock','LY_Rev','LY_Mg%','LY_MgEur']]\
            .rename(columns={'Marca':'Nombre'}),
        on='Nombre', how='outer'
    ).fillna(0)
    detail['Rev vs Budget'] = np.where(
        detail['Budget_Rev'] / 12 > 0,
        detail['Revenue'] / (detail['Budget_Rev'] / 12),
        0
    )
    detail = detail.sort_values('Revenue', ascending=False)
    detail_show = detail[['Nombre','Revenue','Margen_Euros','Margen_Pct',
                           'Budget_Rev','Budget_Mg%','Stock','LY_Rev','LY_Mg%','Rev vs Budget']].copy()
    detail_show.columns = ['Marca','Revenue','Margen€','Mg%','Budget Rev','Budget Mg%',
                           'Stock','LY Rev','LY Mg%','Rev/Budget']

    styled_detail = detail_show.style.format({
        'Revenue':'€{:,.0f}','Margen€':'€{:,.0f}','Mg%':'{:.1%}',
        'Budget Rev':'€{:,.0f}','Budget Mg%':'{:.1%}','Stock':'€{:,.0f}',
        'LY Rev':'€{:,.0f}','LY Mg%':'{:.1%}','Rev/Budget':'{:.1%}'
    }).applymap(
        lambda v: 'color: #2ed573' if isinstance(v,float) and v > 0.2 else
                  ('color: #ff4757' if isinstance(v,float) and v < 0 else ''),
        subset=['Mg%','Budget Mg%','LY Mg%']
    ).applymap(
        lambda v: 'color: #2ed573' if isinstance(v,float) and v >= 1 else
                  ('color: #ff4757' if isinstance(v,float) and 0 < v < 0.8 else ''),
        subset=['Rev/Budget']
    ).set_properties(**{'background-color':'#1a1f2b','color':'#d4dbe8',
                        'border-color':'#252d3d','font-size':'12px'})

    st.dataframe(styled_detail, use_container_width=True, height=380)
